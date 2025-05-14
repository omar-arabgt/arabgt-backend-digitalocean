from django.contrib.auth import views as auth_views
from django.views.generic import ListView, UpdateView, TemplateView, DeleteView, CreateView, FormView
from django.urls import reverse_lazy, reverse
from django.views.generic.detail import DetailView
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils.timezone import localtime
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.utils.timezone import localtime
import openpyxl

from api.models import User, Newsletter, DeletedUser, Group, Forum, Question, Reply, AdminNotification, Post
from api.tasks import send_push_notification, NOTIFICATION_ALL
from .utils import get_merged_user_data, get_signup_method, get_car_sorting_index, CAR_SORTING_LIST
from api.choices import COUNTRIES, GENDERS, STATUS, HOBBIES, INTERESTS, CAR_BRANDS_ITEMS, CAR_SORTING_ITEMS, UserRank
from .forms import *
from django.utils.timezone import make_naive
from django.conf import settings
from django.utils.translation import gettext_lazy as _


class GroupListView(LoginRequiredMixin, ListView):
    """
    Displays a paginated list of groups with optional search functionality.

    Input:
    - Optional query parameter 'q' for searching groups by name.

    Functionality:
    - Retrieves and lists groups, optionally filtered by the search query.

    Output:
    - Renders the 'web/groups/list.html' template with the group list and search context.
    """
    model = Group
    template_name = 'web/groups/list.html'
    context_object_name = 'groups'
    paginate_by = 10

    def get_queryset(self):
        """
        Filters the list of groups based on the search query.
        """
        query = self.request.GET.get('q', '')
        filters = Q()
        if query:
            filters &= Q(name__icontains=query)
        return Group.objects.filter(filters)

    def get_context_data(self, **kwargs):
        """
        Adds additional context data, such as the page name.
        """
        context = super().get_context_data(**kwargs)
        context['page_name'] = 'قائمة المجموعات'
        context['search_query'] = self.request.GET.get('q', '')
        return context


class GroupCreateView(LoginRequiredMixin, CreateView):
    """
    Allows authenticated users to create a new group.

    Input:
    - Form data containing group details.

    Functionality:
    - Creates a new group using the submitted data.

    Output:
    - Redirects to the group list view upon successful creation.
    """
    model = Group
    form_class = GroupForm
    template_name = 'web/groups/create.html'
    success_url = reverse_lazy('group_list')

    def get_context_data(self, **kwargs):
        """
        Adds additional context data, such as the page name.
        """
        context = super().get_context_data(**kwargs)
        context['page_name'] = 'إنشاء مجموعة'
        return context


class GroupUpdateView(LoginRequiredMixin, UpdateView):
    """
    Allows authenticated users to update an existing group.

    Input:
    - Form data containing updated group details.

    Functionality:
    - Updates the specified group with the new data.

    Output:
    - Redirects to the group list view upon successful update.
    """
    model = Group
    form_class = GroupForm
    template_name = 'web/groups/edit.html'
    success_url = reverse_lazy('group_list')

    def get_object(self):
        """
        Retrieves the group object to be updated.
        """
        return get_object_or_404(Group, pk=self.kwargs['pk'])

    def get_context_data(self, **kwargs):
        """
        Adds additional context data, such as the page name.
        """
        context = super().get_context_data(**kwargs)
        context['page_name'] = 'تعديل مجموعة'
        return context


class UserListView(LoginRequiredMixin, ListView):
    """
    Displays a paginated list of users with optional search functionality.

    Input:
    - Optional query parameter 'q' for searching users by username.

    Functionality:
    - Retrieves and lists users, optionally filtered by the search query.

    Output:
    - Renders the 'web/users/list.html' template with the user list and search context.
    """
    model = User
    template_name = 'web/users/list.html'
    context_object_name = 'users'
    paginate_by = 10

    def get_queryset(self):
        """
        Filters the list of users based on the search query and excludes staff and superusers.
        """
        query = self.request.GET.get('q', '')
        filters = Q(is_staff=False, is_superuser=False, emailaddress__verified=True)
        if query:
            filters &= Q(Q(username__icontains=query) | Q(email__icontains=query) | Q(first_name__icontains=query))

        return User.objects.filter(filters)

    def get_context_data(self, **kwargs):
        """
        Adds additional context data, such as the page name.
        """
        context = super().get_context_data(**kwargs)
        context['page_name'] = 'قائمة المستخدمين'
        context['search_query'] = self.request.GET.get('q', '')
        return context


class ViewUserView(LoginRequiredMixin, DetailView):
    """
    Displays the details of a specific user.

    Input:
    - User ID in the URL to retrieve and display user data.

    Functionality:
    - Retrieves the user by ID and displays their details.

    Output:
    - Renders the 'web/users/view.html' template with the user data.
    """
    model = User
    template_name = 'web/users/view.html'
    context_object_name = 'user'

    def get_context_data(self, **kwargs):
        """
        Adds additional context data such as hobbies, interests, car brands, and sorting options.
        """
        context = super().get_context_data(**kwargs)
        context['page_name'] = 'تفاصيل المستخدم'
        context['HOBBIES'] = HOBBIES
        context['INTERESTS'] = INTERESTS
        context['CAR_BRANDS'] = CAR_BRANDS_ITEMS
        context['CAR_SORTING'] = CAR_SORTING_ITEMS
        context['favorite_shows'] = ",".join([str(i) for i in self.object.favorite_shows.all()])
        return context


class ForumListView(LoginRequiredMixin, ListView):
    """
    Displays a paginated list of forums with optional search functionality.

    Input:
    - Optional query parameter 'q' for searching forums by name.

    Functionality:
    - Retrieves and lists forums, optionally filtered by the search query.

    Output:
    - Renders the 'web/forums/list.html' template with the forum list and search context.
    """
    model = Forum
    template_name = 'web/forums/list.html'
    context_object_name = 'forums'
    paginate_by = 10

    def get_queryset(self):
        """
        Filters the list of forums based on the search query.
        """
        query = self.request.GET.get('q', '')
        filters = Q()
        if query:
            filters &= Q(name__icontains=query)
        return Forum.objects.filter(filters)

    def get_context_data(self, **kwargs):
        """
        Adds additional context data, such as the page name.
        """
        context = super().get_context_data(**kwargs)
        context['page_name'] = 'قائمة المنتديات'
        context['search_query'] = self.request.GET.get('q', '')
        return context


class ForumCreateView(LoginRequiredMixin, CreateView):
    """
    Allows authenticated users to create a new forum.

    Input:
    - Form data containing forum details.

    Functionality:
    - Creates a new forum using the submitted data.

    Output:
    - Redirects to the forum list view upon successful creation.
    """
    model = Forum
    form_class = ForumForm
    template_name = 'web/forums/create.html'
    success_url = reverse_lazy('forum_list')

    def get_context_data(self, **kwargs):
        """
        Adds additional context data, such as the page name.
        """
        context = super().get_context_data(**kwargs)
        context['page_name'] = 'انشاء منتدى'
        return context


class ForumUpdateView(LoginRequiredMixin, UpdateView):
    """
    Allows authenticated users to update an existing forum.

    Input:
    - Form data containing updated forum details.

    Functionality:
    - Updates the specified forum with the new data.

    Output:
    - Redirects to the forum list view upon successful update.
    """
    model = Forum
    form_class = ForumForm
    template_name = 'web/forums/edit.html'
    success_url = reverse_lazy('forum_list')

    def get_object(self):
        """
        Retrieves the forum object to be updated.
        """
        return get_object_or_404(Forum, pk=self.kwargs['pk'])

    def get_context_data(self, **kwargs):
        """
        Adds additional context data, such as the page name.
        """
        context = super().get_context_data(**kwargs)
        context['page_name'] = 'تعديل منتدى'
        return context



class ExportUserListView(LoginRequiredMixin, ListView):
    """
    Displays a paginated and filtered list of merged User and DeletedUser data.

    Input:
    - Optional query parameters for filtering the user list by various attributes such as nationality, country, birthdate, etc.

    Functionality:
    - Retrieves and lists merged User and DeletedUser data, optionally filtered by the provided query parameters.

    Output:
    - Renders the 'web/export_users/list.html' template with the merged user list and filtering options.
    """
    template_name = 'web/export_users/list.html'
    context_object_name = 'merged_list'
    paginate_by = 50  # Increased for better performance

    def get_queryset(self):
        """
        Get filtered user data with pagination.
        """
        # Get filter parameters
        query = self.request.GET.get('q', '')
        nationality = self.request.GET.get('nationality')
        country = self.request.GET.get('country')
        birthdate = self.request.GET.get('birthdate')
        gender = self.request.GET.get('gender')
        status = self.request.GET.get('status')
        rank = self.request.GET.get('rank')
        
        # Get filtered data
        return get_merged_user_data(query, nationality, country, birthdate, gender, status, rank)

    def get_context_data(self, **kwargs):
        """
        Add filter parameters to context.
        """
        context = super().get_context_data(**kwargs)
        context['page_name'] = 'Merged User and DeletedUser List'
        context['search_query'] = self.request.GET.get('q', '')
        context['nationality_filter'] = self.request.GET.get('nationality', '')
        context['country_filter'] = self.request.GET.get('country', '')
        context['birthdate_filter'] = self.request.GET.get('birthdate', '')
        context['gender_filter'] = self.request.GET.get('gender', '')
        context['status_filter'] = self.request.GET.get('status', '')
        context['GENDERS'] = GENDERS
        context['STATUS'] = STATUS
        context['COUNTRIES'] = COUNTRIES
        context['RANKS'] = list(UserRank)
        rank = self.request.GET.get('rank', "")
        if rank:
            try:
                context['rank_filter'] = int(rank)
            except ValueError:
                pass
        return context


class ExportUserToExcelView(LoginRequiredMixin, ListView):

    """
    Exports the filtered merged User and DeletedUser data to an Excel file.

    Input:
    - Optional query parameters for filtering the user data by various attributes such as nationality, country, birthdate, etc.

    Functionality:
    - Retrieves and filters the merged user data, writes it to an Excel file, and returns the file as an HTTP response.

    Output:
    - Returns the Excel file containing the merged user data for download.
    """
    def get(self, request, *args, **kwargs):
        """
        Export user data to Excel with streaming response.
        """
        # Get filter parameters
        query = request.GET.get('q', '')
        nationality = request.GET.get('nationality')
        country = request.GET.get('country')
        birthdate = request.GET.get('birthdate')
        gender = request.GET.get('gender')
        status = request.GET.get('status')
        rank = request.GET.get('rank')

        # Create response object for streaming
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=exported_users.xlsx'

        # Create workbook
        wb = openpyxl.Workbook(write_only=True)  # Use write_only mode for better performance
        ws = wb.create_sheet("Users and Deleted Users")

        # Add headers
        headers = [
            'ID',
            'Status',
            'First Name',
            'Last Name',
            'Nick Name',
            'Birth Date',
            'Gender',
            'Nationality',
            'Country',
            'Rank',
            'Delete Reason',
            'Email',
            'Mobile',
            'Authentication',
            'Registration Date',
            'Deletion Date',
            'Newsletter Subscription',
            'Points',
            'Have a Car',
            'Brand Model',
            'Favorite Presenter',
            'Favorite Shows',
            'Hobbies',
            'Is Business Owner',
            *CAR_SORTING_LIST
        ]
        ws.append(headers)

        # Get filtered user data - process active users
        if not status or status == 'active':
            user_filters = Q(is_staff=False, is_superuser=False, emailaddress__verified=True)
            if query:
                user_filters &= Q(Q(username__icontains=query) | Q(email__icontains=query) | 
                                Q(first_name__icontains=query) | Q(last_name__icontains=query) | 
                                Q(nick_name__icontains=query))
            if nationality:
                user_filters &= Q(nationality=nationality)
            if gender:
                user_filters &= Q(gender=gender)
            if country:
                user_filters &= Q(country=country)
            if rank:
                user_filters &= Q(point__gte=rank)
                next_rank_value = UserRank.next_rank_value(int(rank))
                if next_rank_value:
                    user_filters &= Q(point__lt=next_rank_value)
            if birthdate:
                try:
                    date = parse_date(birthdate)
                    if date:
                        user_filters &= Q(birth_date=date)
                except ValueError:
                    pass

            # Process users in batches to reduce memory usage
            batch_size = 500
            users_count = User.objects.filter(user_filters).count()
            
            for offset in range(0, users_count, batch_size):
                users_batch = User.objects.filter(user_filters).order_by('id')[offset:offset+batch_size]
                users_batch = users_batch.select_related('favorite_presenter')

                
                for user in users_batch:
                    # Convert timezone-aware datetime to naive datetime
                    registration_date = make_naive(user.date_joined).strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else ''
                    
                    # Get favorite shows as string
                    favorite_shows = ",".join([str(i) for i in user.favorite_shows]) if isinstance(user.favorite_shows, list) else ''
                    
                    # Get hobbies as string
                    hobbies = ', '.join(user.hobbies) if user.hobbies else ''
                    
                    # Get car sorting values
                    car_sorting_values = get_car_sorting_index(user.car_sorting)
                    
                    # Prepare row data
                    row = [
                        user.id,
                        "مفعل",
                        user.first_name,
                        user.last_name,
                        user.nick_name,
                        user.birth_date.strftime('%Y-%m-%d') if user.birth_date else '',
                        "ذكر" if user.gender == "M" else "انثي" if user.gender == "F" else "",
                        user.get_nationality_display(),
                        user.get_country_display(),
                        user.rank,
                        "",  # Delete reason (empty for active users)
                        user.email,
                        user.phone_number,
                        get_signup_method(user.id),
                        registration_date,
                        "",  # Deletion date (empty for active users)
                        'Yes' if getattr(user, 'send_notification', False) else 'No',
                        user.point,
                        'Yes' if user.has_car else 'No',
                        user.car_type,
                        str(user.favorite_presenter) if user.favorite_presenter else '',
                        favorite_shows,
                        hobbies,
                        'Yes' if user.has_business else 'No',
                        *car_sorting_values
                    ]
                    
                    ws.append(row)

        # Process deleted users
        if not status or status == 'deleted':
            deleted_user_filters = Q()
            if query:
                deleted_user_filters &= Q(Q(username__icontains=query) | Q(email__icontains=query) | 
                                       Q(first_name__icontains=query) | Q(last_name__icontains=query) | 
                                       Q(nick_name__icontains=query))
            if nationality:
                deleted_user_filters &= Q(nationality=nationality)
            if gender:
                deleted_user_filters &= Q(gender=gender)
            if country:
                deleted_user_filters &= Q(country=country)
            if rank:
                deleted_user_filters &= Q(point__gte=rank)
                next_rank_value = UserRank.next_rank_value(int(rank))
                if next_rank_value:
                    deleted_user_filters &= Q(point__lt=next_rank_value)
            if birthdate:
                try:
                    date = parse_date(birthdate)
                    if date:
                        deleted_user_filters &= Q(birth_date=date)
                except ValueError:
                    pass

            # Process deleted users in batches
            batch_size = 500
            deleted_count = DeletedUser.objects.filter(deleted_user_filters).count()
            
            for offset in range(0, deleted_count, batch_size):
                deleted_batch = DeletedUser.objects.filter(deleted_user_filters).order_by('id')[offset:offset+batch_size]
                deleted_batch = deleted_batch.select_related('favorite_presenter')
                
                for user in deleted_batch:
                    # Process similar to active users but with deleted user data
                    deletion_date = make_naive(user.last_login).strftime('%Y-%m-%d %H:%M:%S') if getattr(user, 'last_login', None) else ''
                    
                    favorite_shows = ",".join([str(i) for i in user.favorite_shows]) if isinstance(user.favorite_shows, list) else ''
                    hobbies = ', '.join(user.hobbies) if user.hobbies else ''
                    car_sorting_values = get_car_sorting_index(user.car_sorting)
                    
                    row = [
                        user.user_id,
                        "محذوف",
                        user.first_name,
                        user.last_name,
                        user.nick_name,
                        user.birth_date.strftime('%Y-%m-%d') if user.birth_date else '',
                        "ذكر" if user.gender == "M" else "انثي" if user.gender == "F" else "",
                        user.get_nationality_display(),
                        user.get_country_display(),
                        user.rank,
                        user.delete_reason,
                        user.email,
                        user.phone_number,
                        get_signup_method(user.user_id),
                        "",  # Registration date not available for deleted users
                        deletion_date,
                        'Yes' if getattr(user, 'send_notification', False) else 'No',
                        user.point,
                        'Yes' if user.has_car else 'No',
                        user.car_type,
                        str(user.favorite_presenter) if user.favorite_presenter else '',
                        favorite_shows,
                        hobbies,
                        'Yes' if user.has_business else 'No',
                        *car_sorting_values
                    ]
                    
                    ws.append(row)

        # Save workbook to response
        wb.save(response)
        return response
class NewsletterListView(LoginRequiredMixin, ListView):
    """
    Displays a paginated list of newsletter subscriptions with optional search functionality.

    Input:
    - Optional query parameter 'q' for searching newsletter subscriptions by email.

    Functionality:
    - Retrieves and lists newsletter subscriptions, optionally filtered by the search query.

    Output:
    - Renders the 'web/newsletter/list.html' template with the newsletter list and search context.
    """
    model = Newsletter
    template_name = 'web/newsletter/list.html'
    context_object_name = 'newsletter_list'
    paginate_by = 10

    def get_queryset(self):
        """
        Filters the newsletter list based on the search query.
        """
        query = self.request.GET.get('q')
        if query:
            return Newsletter.objects.filter(Q(email__icontains=query))
        return Newsletter.objects.all()

    def get_context_data(self, **kwargs):
        """
        Adds additional context data, such as the page name and search query.
        """
        context = super().get_context_data(**kwargs)
        context['page_name'] = 'قائمة النشرة البريدية'
        context['search_query'] = self.request.GET.get('q', '')
        return context


class UserDeleteView(LoginRequiredMixin, DeleteView):
    """
    Delete a specific user's information.

    Input:
    - User ID is passed in the URL.

    Functionality:
    - Retrieves the user by ID and deletes their information.

    Output:
    - Redirects to the user list view upon successful deletion.
    """
    model = User
    success_url = reverse_lazy('user_list')
    
    def get_object(self, queryset=None):
        """
        Retrieves the user object to be deleted.
        """
        return get_object_or_404(User, pk=self.kwargs['pk'])

    def post(self, request, *args, **kwargs):
        """
        Handles the deletion of the user, ensuring that superusers and staff members cannot be deleted.
        """
        user = self.get_object()
        delete_reason = request.POST.get('delete_reason', '').strip()

        
        if user.is_superuser or user.is_staff:
            return redirect(self.success_url)

        user.delete(delete_reason=delete_reason)
        return redirect(self.success_url)


class DeletedUserListView(LoginRequiredMixin, ListView):
    """
    Displays a paginated list of deleted users with optional search and filter functionality.

    Input:
    - Optional query parameters for searching and filtering deleted users by various attributes such as username, nationality, country, birthdate, etc.

    Functionality:
    - Retrieves and lists deleted users, optionally filtered by the search and filter criteria.

    Output:
    - Renders the 'web/deleted_users/list.html' template with the deleted user list and filter options.
    """
    model = DeletedUser
    template_name = 'web/deleted_users/list.html'
    context_object_name = 'deleted_user_list'
    paginate_by = 10

    def get_queryset(self):
        """
        Filters the list of deleted users based on the search and filter criteria.
        """
        query = self.request.GET.get('q', '')
        nationality = self.request.GET.get('nationality')
        country = self.request.GET.get('country')
        birthdate = self.request.GET.get('birthdate')

        filters = Q()

        if query:
            filters &= Q(Q(username__icontains=query) | Q(email__icontains=query) | Q(first_name__icontains=query) | Q(last_name__icontains=query) | Q(nick_name__icontains=query))
        if nationality:
            filters &= Q(nationality=nationality)
        if country:
            filters &= Q(country=country)
        if birthdate:
            try:
                date = parse_date(birthdate)
                if date:
                    filters &= Q(birth_date__gt=date)
            except ValueError:
                pass

        return DeletedUser.objects.filter(filters)

    def get_context_data(self, **kwargs):
        """
        Adds additional context data, such as filter options and the page name.
        """
        context = super().get_context_data(**kwargs)
        context['page_name'] = 'قائمة المستخدمين المحذوفين'
        context['search_query'] = self.request.GET.get('q', '')
        context['nationality_filter'] = self.request.GET.get('nationality', '')
        context['country_filter'] = self.request.GET.get('country', '')
        context['birthdate_filter'] = self.request.GET.get('birthdate', '')
        context['car_brand_filter'] = self.request.GET.get('car_brand', '')
        context['COUNTRIES'] = COUNTRIES
        return context


def download_newsletter_excel(request):
    """
    Downloads an Excel file containing the newsletter subscriptions.

    Input:
    - No specific input required.

    Functionality:
    - Retrieves all newsletter subscriptions and writes them to an Excel file.

    Output:
    - Returns the Excel file as an HTTP response for download.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Newsletter"

    headers = ['ID', 'Email', 'Created At']
    ws.append(headers)

    newsletters = Newsletter.objects.all().values_list('id', 'email', 'created_at')
    for newsletter in newsletters:
        id, email, created_at = newsletter
        if created_at:
            created_at = localtime(created_at).replace(tzinfo=None)
        ws.append([id, email, created_at])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=newsletter.xlsx'

    wb.save(response)
    return response


class LoginView(auth_views.LoginView):
    """
    Handles user login.

    Input:
    - Form data containing username and password.

    Functionality:
    - Authenticates and logs in the user.

    Output:
    - Renders the 'web/login.html' template with the login form.
    """
    template_name = "web/login.html"
    form_class = CustomAuthenticationForm

    def get_context_data(self, **kwargs):
        """
        Adds additional context data, such as the page name.
        """
        context = super().get_context_data(**kwargs)
        context['page_name'] = 'تسجيل الدخول'
        return context


class HomeView(LoginRequiredMixin, TemplateView):
    """
    Displays the home/dashboard page for logged-in users.

    Input:
    - No specific input required.

    Functionality:
    - Renders the home page for authenticated users.

    Output:
    - Renders the 'web/home.html' template.
    """
    template_name = "web/home.html"

    def get_context_data(self, **kwargs):
        """
        Adds additional context data, such as the page name.
        """
        context = super().get_context_data(**kwargs)
        context['page_name'] = 'لوحة التحكم'
        return context


class TermsOfUsePrivacyPolicy(TemplateView):
    """
    Displays the Terms of Use and Privacy Policy page.

    Input:
    - No specific input required.

    Functionality:
    - Renders the Terms of Use and Privacy Policy page.

    Output:
    - Renders the 'web/terms_of_us_privacy_policy.html' template.
    """
    template_name = 'web/terms_of_us_privacy_policy.html'

    def get_context_data(self, **kwargs):
        """
        Adds additional context data, such as a flag to hide the sidebar.
        """
        context = super().get_context_data(**kwargs)
        context['no_sidebar'] = True
        return context

class NotificationView(LoginRequiredMixin, FormView):
    """
    Displays the home/notifications page for logged-in users.
    Input:
    - No specific input required.
    Functionality:
    - Renders the notification page for authenticated users.
    - Handles the submission of the notification form.
    - Paginates sent notifications.

    Output:
    - Renders the 'web/notifications/container.html' template with the notification form and list of sent notifications.
    """

    template_name = 'web/notifications/container.html'
    form_class = NotificationForm
    paginate_by = 10
    success_url = '?tab=2&page=1'

    def get_context_data(self, **kwargs):
        """
        Adds additional context data, such as the page name, form, and paginated notifications.
        """
        context = super().get_context_data(**kwargs)
        context['page_name'] = 'إرسال التنبيهات'
        notifications = AdminNotification.objects.all().order_by('-created_at')
        paginator = Paginator(notifications, self.paginate_by)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['page_obj'] = page_obj
        return context
  
    def form_valid(self, form):
        title = form.cleaned_data.get("title")
        content = form.cleaned_data.get("content")
        external_link = form.cleaned_data.get("link")
        post_id = form.cleaned_data.get("post_id")
        link = None

        if post_id:
            external_link = None
            exists = Post.objects.filter(post_id=post_id).exists()
            if exists:
                link = f"{settings.APP_URL}/post-details?id={post_id}"
            else:
                form.add_error("post_id", _("Post does not exist!"))
                return self.form_invalid(form)

        send_push_notification.delay(NOTIFICATION_ALL, title, content, link, external_link, True)
        return super().form_valid(form)
  
    def post(self, request, *args, **kwargs):
        """
        Handles the submission of the notification form and sends notifications to all users.
        """
        form = self.get_form()

        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

class ForumGroupQuestionsView(LoginRequiredMixin, TemplateView):
    """
    Displays the home/questions page for logged-in users.

    Input:
    - No specific input required.

    Functionality:
    - Renders the questions for (group & fourms) page for authenticated users.

    Output:
    - Renders the 'web/questions.html' template.
    """

    template_name = "web/questions/questions.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tab = self.request.GET.get('tab', 'forums')
        extra_query_params = any(key != 'tab' for key in self.request.GET.keys())

        if tab == 'forums':
            questions = self.get_filtered_questions(tab)
            paginator = Paginator(questions, 10)
            page_number = self.request.GET.get('page')
            context['page_obj'] = paginator.get_page(page_number)
            context['forums_questions'] = context['page_obj']
        elif tab == 'groups':
            questions = self.get_filtered_questions(tab)
            paginator = Paginator(questions, 10)
            page_number = self.request.GET.get('page')
            context['page_obj'] = paginator.get_page(page_number)
            context['groups_questions'] = context['page_obj']

        context['active_tab'] = tab
        context[f'{tab}_search_query'] = self.request.GET.get(f'{tab}_q', '')
        context[f'{tab}_question_id'] = self.request.GET.get(f'{tab}_question_id', '')
        context['forum_id'] = int(self.request.GET.get('forum_id')) if self.request.GET.get('forum_id', False) else ""
        context['group_id'] = int(self.request.GET.get('group_id')) if self.request.GET.get('group_id', False) else ""
        context[f'{tab}_date'] = self.request.GET.get(f'{tab}_date', '')
        context['forums'] = Forum.objects.all()
        context['groups'] = Group.objects.all()
        context['page_name'] = 'الأسئلة و النقاشات'
        context['show_clear_button'] = extra_query_params

        return context

    def get_filtered_questions(self, tab):
        if tab == 'forums':
            questions = Question.objects.filter(forum__isnull=False)
            search_query = self.request.GET.get(f'{tab}_q', '')
            question_id = self.request.GET.get(f'{tab}_question_id', '')
            forum_id = self.request.GET.get('forum_id', '')
            date_filter = self.request.GET.get(f'{tab}_date', '')

            if search_query:
                questions = questions.filter(
                    Q(user__username__icontains=search_query) | Q(user__id__icontains=search_query)
                )
            if question_id:
                questions = questions.filter(id=question_id)
            if forum_id:
                questions = questions.filter(forum__id=forum_id)
            if date_filter:
                questions = questions.filter(created_at__date=date_filter)

        elif tab == 'groups':
            questions = Question.objects.filter(group__isnull=False)
            search_query = self.request.GET.get(f'{tab}_q', '')
            question_id = self.request.GET.get(f'{tab}_question_id', '')
            group_id = self.request.GET.get('group_id', '')
            date_filter = self.request.GET.get(f'{tab}_date', '')

            if search_query:
                questions = questions.filter(
                    Q(user__username__icontains=search_query) | Q(user__id__icontains=search_query)
                )
            if question_id:
                questions = questions.filter(id=question_id)
            if group_id:
                questions = questions.filter(group__id=group_id)
            if date_filter:
                questions = questions.filter(created_at__date=date_filter)

        return questions

class QuestionDetailView(LoginRequiredMixin, DetailView):
    model = Question
    template_name = 'web/questions/question_detail.html'
    context_object_name = 'question'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        question = self.get_object()
        
        replies = question.replies.filter(parent_reply__isnull=True)
        
        context['replies'] = self.annotate_replies_with_likes(replies)
        context['no_sidebar'] = True
        context['page_name'] = 'تفاصيل السؤال'
        context['total_comments'] = count_replies(question)

        return context

    def annotate_replies_with_likes(self, replies):
        for reply in replies:
            reply.total_likes = reply.reactions.filter(reaction_type='like').count()
            if reply.replies.exists():
                reply.nested_replies = self.annotate_replies_with_likes(reply.replies.all())
        return replies

#  for getting nested reply count
def count_replies(question):
    total_replies = question.replies.count()
    for reply in question.replies.all():
        total_replies += count_nested_replies(reply)
    return total_replies

def count_nested_replies(reply):
    total_replies = reply.replies.count()
    for nested_reply in reply.replies.all():
        total_replies += count_nested_replies(nested_reply)
    return total_replies

@login_required
def delete_reply(request, reply_id):
    reply = get_object_or_404(Reply, id=reply_id)
    question_id = reply.question.id if reply.question else reply.parent_reply.question.id

    if request.method == 'POST':
        reply.delete()
        return redirect(reverse('question_detail', args=[question_id]))

    return redirect(reverse('question_detail', args=[question_id]))

@login_required
def delete_question(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    
    if request.method == 'POST':
        question.delete()
        return redirect(reverse('questions'))
    
    return redirect(reverse('question_detail', args=[question_id]))

from api.models import Post
from .forms import PostManageForm
from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist

class PostManageView(FormView):
    template_name = 'web/posts_manage/container.html'
    form_class = PostManageForm
    success_url = '?page=1'

    def form_valid(self, form):
        post_id = form.cleaned_data['post_id']
        action = form.cleaned_data['action']
        
        try:
            post = Post.objects.get(post_id=post_id)
            
            if action == 'draft':
                post.is_published = False
                post.save()
                messages.success(self.request, "تم تحويل المقال إلى مسودة بنجاح.")
            elif action == 'publish':
                post.is_published = True
                post.save()
                messages.success(self.request, "تم نشر المقال بنجاح.")

            elif action == 'delete':
                post.delete()
                messages.success(self.request, "تم حذف المقال بنجاح.")
        except ObjectDoesNotExist:
            messages.error(self.request, "لم يتم العثور على مقال بالمعرف المحدد.")
        
        return super().form_valid(form)